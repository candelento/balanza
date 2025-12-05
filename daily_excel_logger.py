import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import logging
from typing import List, Any, Optional, Dict

# Configure logging for this module
# Create a logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO) # Set minimum logging level

# Create handlers
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO) # Console logs INFO and above

# Create a file handler
log_file_path = "api.log" # Define log file path
file_handler = logging.FileHandler(log_file_path)
file_handler.setLevel(logging.DEBUG) # File logs DEBUG and above

# Create formatters and add them to handlers
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

# Add handlers to the logger
if not logger.handlers: # Avoid adding handlers multiple times if the module is reloaded
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

# Configuration
EXCEL_FILENAME = "daily_log.xlsx"
HEADERS = [
    "Registro ID", "Tipo Operación", "Contraparte", "Producto",
    "Peso Bruto (kg)", "Peso Tara (kg)", "Merma (kg)", "Peso Neto (kg)",
    "Precio x Kg", "Importe", "Chofer/Transporte", "Patente", "Incoterm",
    "Fecha Operacion", "Hora Ingreso", "Hora Salida", "Remito", "Observaciones"
]
# Column indices (1-based)
ID_COLUMN_INDEX = 1
TYPE_COLUMN_INDEX = 2 # Index for "Tipo Operación"
OBSERVACIONES_COLUMN_INDEX = 18 # Index for "Observaciones" (moved due to new 'Incoterm' and 'Remito' columns)

# --- Styling ---
HEADER_FONT = Font(bold=True, color="FFFFFF") # White text
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue background
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
COLUMN_WIDTHS = {
    'A': 12,  # Registro ID
    'B': 15,  # Tipo Operación
    'C': 25,  # Contraparte
    'D': 25,  # Producto
    'E': 18,  # Peso Bruto (kg)
    'F': 18,  # Peso Tara (kg)
    'G': 12,  # Merma (kg)
    'H': 18,  # Peso Neto (kg)
    'I': 15,  # Precio x Kg
    'J': 18,  # Importe
    'K': 18,  # Chofer/Transporte
    'L': 15,  # Patente
    'M': 10,  # Incoterm
    'N': 15,  # Fecha Operacion
    'O': 15,  # Hora Ingreso
    'P': 15,  # Hora Salida
    'Q': 14,  # Remito
    'R': 30   # Observaciones
}

def _apply_sheet_formatting(sheet):
    """Applies formatting (column widths, header style) to the sheet."""
    if sheet.max_row == 0: return

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT

    for col_letter, width in COLUMN_WIDTHS.items():
         try:
             sheet.column_dimensions[col_letter].width = width
         except Exception as e:
             logger.warning(f"Could not set width for column {col_letter}: {e}")

    sheet.freeze_panes = 'A2'
    logger.info(f"Applied formatting to sheet '{sheet.title}'")

# --- Workbook and Sheet Management ---

def _load_or_create_workbook(filename):
    """Loads an existing workbook or creates a new one."""
    logger.debug(f"Attempting to load or create workbook: {filename}")
    try:
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
            logger.info(f"Loaded existing workbook: {filename}")
        else:
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            logger.info(f"Created new workbook: {filename}")
    except InvalidFileException:
        logger.error(f"Error reading {filename}. Creating a new one.", exc_info=True)
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
    except Exception as e:
        logger.error(f"Unexpected error loading/creating workbook {filename}: {e}", exc_info=True)
        workbook = Workbook() # Fallback
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
    return workbook

def _get_or_create_daily_sheet(workbook):
    """Gets or creates the sheet for the current day (YYYY-MM-DD)."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    logger.debug(f"Getting or creating sheet for today: {today_str}")
    if today_str not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=today_str)
        sheet.append(HEADERS)
        _apply_sheet_formatting(sheet)
        logger.info(f"Created new sheet for {today_str} and applied formatting.")
    else:
        sheet = workbook[today_str]
        logger.debug(f"Using existing sheet for today: {today_str}")
        # MIGRATION: ensure 'Incoterm' column exists after 'Patente'
        try:
            existing_headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            if existing_headers and "Incoterm" not in existing_headers:
                # Find position of 'Patente' header (1-based index); default to 12 (column L)
                try:
                    patente_idx = existing_headers.index("Patente") + 1
                except ValueError:
                    patente_idx = 12
                insert_at = patente_idx + 1
                sheet.insert_cols(insert_at)
                sheet.cell(row=1, column=insert_at, value="Incoterm")
                logger.info(f"Inserted missing 'Incoterm' column at position {insert_at} in sheet '{sheet.title}'.")
                # Re-apply header styling and widths (best-effort)
                _apply_sheet_formatting(sheet)
        except Exception as e:
            logger.warning(f"Could not migrate sheet '{sheet.title}' to include 'Incoterm' column: {e}")
        # MIGRATION: ensure 'Remito' column exists before 'Observaciones'
        try:
            existing_headers = [cell.value for cell in sheet[1]] if sheet.max_row >= 1 else []
            if existing_headers and "Remito" not in existing_headers:
                # Find position of 'Observaciones' header (1-based index); default to len(existing_headers)+1
                try:
                    obs_idx = existing_headers.index("Observaciones") + 1
                except ValueError:
                    obs_idx = len(existing_headers) + 1
                insert_at = obs_idx  # insert before Observaciones
                sheet.insert_cols(insert_at)
                sheet.cell(row=1, column=insert_at, value="Remito")
                logger.info(f"Inserted missing 'Remito' column at position {insert_at} in sheet '{sheet.title}'.")
                _apply_sheet_formatting(sheet)
        except Exception as e:
            logger.warning(f"Could not migrate sheet '{sheet.title}' to include 'Remito' column: {e}")
    return sheet

def _save_workbook(workbook, filename):
    """Saves the workbook."""
    logger.debug(f"Attempting to save workbook: {filename}")
    try:
        workbook.save(filename)
        logger.info(f"Workbook saved successfully: {filename}")
    except Exception as e:
        logger.error(f"Error saving workbook '{filename}': {e}", exc_info=True)

def _find_row_by_id_and_type(sheet, entry_id, entry_type):
    """Finds the row index for a given ID and Type within the specific sheet."""
    logger.debug(f"Searching for row with ID: {entry_id}, Type: {entry_type} in sheet: {sheet.title}")
    if sheet.max_row <= 1:
        logger.debug("Sheet is empty or only has headers. Row not found.")
        return None

    # Normalize search entry_type for robust comparison
    search_type = entry_type.strip() if isinstance(entry_type, str) else entry_type

    for row_idx in range(2, sheet.max_row + 1):
        id_cell_value = sheet.cell(row=row_idx, column=ID_COLUMN_INDEX).value
        type_cell_value = sheet.cell(row=row_idx, column=TYPE_COLUMN_INDEX).value

        # Normalize sheet type value for robust comparison
        sheet_type = type_cell_value.strip() if isinstance(type_cell_value, str) else type_cell_value

        # Check if type matches first
        if sheet_type != search_type:
            continue

        logger.debug(f"Comparing: Sheet ID='{id_cell_value}' (Type: {type(id_cell_value)}), Sheet Type='{type_cell_value}' (Type: {type(type_cell_value)}) with Search ID='{entry_id}' (Type: {type(entry_id)}), Search Type='{entry_type}' (Type: {type(entry_type)})")

        # Then check ID (converting to int for comparison)
        try:
            # Attempt strict integer comparison first
            if id_cell_value is not None and isinstance(id_cell_value, (int, float)) and int(id_cell_value) == int(entry_id):
                logger.debug(f"Match found by strict integer comparison at row {row_idx}")
                return row_idx
        except (ValueError, TypeError):
            logger.debug(f"Strict integer comparison failed for ID '{id_cell_value}' vs '{entry_id}'. Trying string comparison.")
            # Fallback to string comparison if integer conversion/comparison fails
            try:
                if str(id_cell_value).strip() == str(entry_id).strip(): # Use strip to handle potential spaces
                     logger.debug(f"Match found by string comparison at row {row_idx}")
                     return row_idx
            except Exception as str_e:
                 logger.warning(f"Error during string comparison for ID '{id_cell_value}' vs '{entry_id}': {str_e}")

    logger.debug(f"Row with ID: {entry_id}, Type: {entry_type} not found in sheet: {sheet.title}")
    return None # Not found

# --- Public Data Management Functions ---

def upsert_data(entry_id: int, entry_type: str, data_row: List[Any], filename=EXCEL_FILENAME):
    """
    Updates or inserts data in the current day's sheet based on the numeric entry_id and entry_type.
    """
    logger.info(f"Attempting to upsert data for ID: {entry_id}, Type: {entry_type}")
    if entry_id is None:
         logger.error("Cannot upsert data without a valid entry_id.")
         return
    if not entry_type:
         logger.error("Cannot upsert data without a valid entry_type.")
         return

    if len(data_row) != len(HEADERS):
         logger.error(f"Data row length ({len(data_row)}) != headers length ({len(HEADERS)}). ID: {entry_id}, Type: {entry_type}")
         return

    try:
        workbook = _load_or_create_workbook(filename)
        if workbook is None: return

        sheet = _get_or_create_daily_sheet(workbook)

        # Find row using both ID and Type
        row_idx_to_update = _find_row_by_id_and_type(sheet, entry_id, entry_type) 

        if row_idx_to_update:
            logger.debug(f"Found existing row for ID {entry_id} (Type: {entry_type}) at row {row_idx_to_update}. Updating.")
            for col_idx, value in enumerate(data_row, start=1):
                # Check if it's a numeric column and value is None or empty string
                # Use 0-based index for checking against numeric column indices
                if (col_idx - 1) in [4, 5, 6, 7] and (value is None or value == ""):
                    cell_value_to_write = None # Write None for empty numeric fields
                else:
                    cell_value_to_write = value # Otherwise, use the provided value

                sheet.cell(row=row_idx_to_update, column=col_idx, value=cell_value_to_write)
            logger.info(f"Updated row for ID {entry_id} (Type: {entry_type}) in sheet {sheet.title}")
        else:
            logger.debug(f"No existing row found for ID {entry_id} (Type: {entry_type}). Appending new row.")
            # When appending, also handle numeric None/empty strings
            cleaned_data_row = []
            for i, value in enumerate(data_row):
                 if i in [4, 5, 6, 7] and (value is None or value == ""):
                      cleaned_data_row.append(None)
                 else:
                      cleaned_data_row.append(value)
            sheet.append(cleaned_data_row)
            logger.info(f"Appended new row for ID {entry_id} (Type: {entry_type}) to sheet {sheet.title}")

        _save_workbook(workbook, filename)

    except Exception as e:
        logger.error(f"Failed to upsert data for ID {entry_id} (Type: {entry_type}): {e}", exc_info=True)


def delete_data(entry_id: int, entry_type: str, filename=EXCEL_FILENAME):
    """
    Deletes the row corresponding to the given numeric entry_id and entry_type from the current day's sheet.
    """
    logger.info(f"Attempting to delete data for ID: {entry_id}, Type: {entry_type}")
    if entry_id is None:
         logger.error("Cannot delete data without a valid entry_id.")
         return
    if not entry_type:
         logger.error("Cannot delete data without a valid entry_type.")
         return

    try:
        workbook = _load_or_create_workbook(filename)
        if workbook is None: return

        today_str = datetime.now().strftime("%Y-%m-%d")
        if today_str not in workbook.sheetnames:
             logger.warning(f"Sheet for today '{today_str}' not found. Cannot delete ID {entry_id} (Type: {entry_type}).")
             return

        sheet = workbook[today_str]
        # Find row using both ID and Type
        row_idx_to_delete = _find_row_by_id_and_type(sheet, entry_id, entry_type) 

        if row_idx_to_delete:
            logger.debug(f"Found row to delete for ID {entry_id} (Type: {entry_type}) at row {row_idx_to_delete}.")
            sheet.delete_rows(row_idx_to_delete)
            logger.info(f"Deleted row for ID {entry_id} (Type: {entry_type}) from sheet {sheet.title}")
            _save_workbook(workbook, filename)
        else:
            logger.warning(f"Could not find row with ID {entry_id} (Type: {entry_type}) in sheet {sheet.title} to delete.")

    except Exception as e:
        logger.error(f"Failed to delete data for ID {entry_id} (Type: {entry_type}): {e}", exc_info=True)

# --- Data Loading Function ---

def load_data_by_date(date_str: str, filename=EXCEL_FILENAME) -> Dict[str, List[Dict[str, Any]]]:
    """
    Loads data from a specific sheet identified by date_str (YYYY-MM-DD)
    into a dictionary of lists, separated by entry type ('Compra' or 'Venta').
    """
    logger.info(f"Attempting to load data for date: {date_str} from: {filename}")
    data = {"Compra": [], "Venta": []}
    try:
        workbook = _load_or_create_workbook(filename)
        if workbook is None:
            logger.warning("Workbook could not be loaded or created. Returning empty data.")
            return data

        if date_str not in workbook.sheetnames:
            logger.info(f"No sheet found for date '{date_str}'. Returning empty data.")
            return data

        sheet = workbook[date_str]
        headers = [cell.value for cell in sheet[1]]

        if not headers:
            logger.warning(f"Sheet '{sheet.title}' is empty or has no headers.")
            return data

        header_keys = []
        for header in headers:
            key = header.lower().replace(" ", "_").replace("(kg)", "").strip()
            if header.endswith("(kg)") and key.endswith("_"):
                key = key[:-1]
            header_keys.append(key)
        
        logger.debug(f"Sheet headers: {headers}")
        logger.debug(f"Mapped header keys: {header_keys}")

        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, cell in enumerate(sheet[row_idx]):
                if col_idx < len(header_keys):
                    header_key = header_keys[col_idx]
                    cell_value = cell.value
                    
                    if header_key in ["peso_bruto", "peso_tara", "merma", "peso_neto", "precio_x_kg", "importe"]:
                        try:
                            cell_value = float(cell_value) if cell_value is not None else None
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert value '{cell_value}' in column '{headers[col_idx]}' to float in sheet '{sheet.title}', row {row_idx}. Setting to None.")
                            cell_value = None
                            
                    row_data[header_key] = cell_value
            
            entry_type = row_data.get("tipo_operación")
            if entry_type == "Compra":
                compra_entry = {
                    "id": row_data.get("registro_id"),
                    "proveedor": row_data.get("contraparte"),
                    "mercaderia": row_data.get("producto"),
                    "bruto": row_data.get("peso_bruto"),
                    "tara": row_data.get("peso_tara"),
                    "merma": row_data.get("merma"),
                    "neto": row_data.get("peso_neto"),
                    "precio_kg": row_data.get("precio_x_kg"),
                    "importe": row_data.get("importe"),
                    "chofer": row_data.get("chofer/transporte"),
                    "patente": row_data.get("patente"),
                    "fecha": row_data.get("fecha_operacion"),
                    "hora_ingreso": row_data.get("hora_ingreso"),
                    "hora_salida": row_data.get("hora_salida"),
                    "observaciones": row_data.get("observaciones")
                }
                data["Compra"].append(compra_entry)
            elif entry_type == "Venta":
                venta_entry = {
                    "id": row_data.get("registro_id"),
                    "cliente": row_data.get("contraparte"),
                    "mercaderia": row_data.get("producto"),
                    "bruto": row_data.get("peso_bruto"),
                    "tara": row_data.get("peso_tara"),
                    "merma": row_data.get("merma"),
                    "neto": row_data.get("peso_neto"),
                    "precio_kg": row_data.get("precio_x_kg"),
                    "importe": row_data.get("importe"),
                    "transporte": row_data.get("chofer/transporte"),
                    "patente": row_data.get("patente"),
                    "incoterm": row_data.get("incoterm"),
                    "fecha": row_data.get("fecha_operacion"),
                    "hora_ingreso": row_data.get("hora_ingreso"),
                    "hora_salida": row_data.get("hora_salida"),
                    "remito": row_data.get("remito"),
                    "observaciones": row_data.get("observaciones")
                }
                data["Venta"].append(venta_entry)

        logger.info(f"Successfully loaded {len(data['Compra'])} Compra entries and {len(data['Venta'])} Venta entries for date {date_str}.")

    except Exception as e:
        logger.error(f"Error loading data for date {date_str} from Excel: {e}", exc_info=True)

    return data


def load_daily_data(filename=EXCEL_FILENAME) -> Dict[str, List[Dict[str, Any]]]:
    """
    Loads data from the current day's sheet into a dictionary of lists,
    separated by entry type ('Compra' or 'Venta').
    """
    logger.info(f"Attempting to load daily data from: {filename}")
    data = {"Compra": [], "Venta": []}
    try:
        workbook = _load_or_create_workbook(filename)
        if workbook is None:
            logger.warning("Workbook could not be loaded or created. Returning empty data.")
            return data

        today_str = datetime.now().strftime("%Y-%m-%d")
        if today_str not in workbook.sheetnames:
            logger.info(f"No sheet found for today '{today_str}'. Returning empty data.")
            return data

        sheet = workbook[today_str]
        headers = [cell.value for cell in sheet[1]] # Read headers from the sheet

        if not headers:
            logger.warning(f"Sheet '{sheet.title}' is empty or has no headers.")
            return data

        # Map header names to dictionary keys (using lowercase and replacing spaces with underscores)
        # Correctly handle keys for numeric columns by removing trailing underscore if present
        header_keys = []
        for header in headers:
            key = header.lower().replace(" ", "_").replace("(kg)", "").strip()
            # Remove trailing underscore if the original header ended with "(kg)"
            if header.endswith("(kg)") and key.endswith("_"):
                key = key[:-1]
            header_keys.append(key)
        logger.debug(f"Sheet headers: {headers}")
        logger.debug(f"Mapped header keys: {header_keys}")


        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, cell in enumerate(sheet[row_idx]):
                if col_idx < len(header_keys):  
                    header_key = header_keys[col_idx]
                    cell_value = cell.value
                    
                    # Attempt to convert numeric fields to float
                    if header_key in ["peso_bruto", "peso_tara", "merma", "peso_neto", "precio_x_kg", "importe"]:
                        original_value = cell_value # Store original value for logging
                        try:
                            cell_value = float(cell_value) if cell_value is not None else None
                            if original_value is not None and cell_value is None:
                                # Log if conversion failed for a non-None value
                                logger.warning(f"Conversion to float failed for value '{original_value}' in column '{headers[col_idx]}' (key: {header_key}) in sheet '{sheet.title}', row {row_idx}. Setting to None.")
                            elif original_value != cell_value:
                                # Log if conversion changed the value (e.g., from int to float)
                                logger.debug(f"Converted value '{original_value}' to float '{cell_value}' for column '{headers[col_idx]}' (key: {header_key}) in sheet '{sheet.title}', row {row_idx}.")
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert value '{original_value}' in column '{headers[col_idx]}' (key: {header_key}) to float in sheet '{sheet.title}', row {row_idx}. Setting to None.")
                            cell_value = None # Set to None if conversion fails
                            
                    row_data[header_key] = cell_value
            logger.debug(f"Read row {row_idx}: {row_data}")

            # Determine entry type and add to appropriate list
            entry_type = row_data.get("tipo_operación")
            if entry_type == "Compra":
                # Map Excel data to Compra model keys
                compra_entry = {
                    "id": row_data.get("registro_id"),
                    "proveedor": row_data.get("contraparte"),
                    "mercaderia": row_data.get("producto"),
                    "bruto": row_data.get("peso_bruto"),
                    "tara": row_data.get("peso_tara"),
                    "merma": row_data.get("merma"),
                    "neto": row_data.get("peso_neto"),
                    "precio_kg": row_data.get("precio_x_kg"),
                    "importe": row_data.get("importe"),
                    "chofer": row_data.get("chofer/transporte"),
                    "patente": row_data.get("patente"),
                    "fecha": row_data.get("fecha_operacion"),
                    "hora_ingreso": row_data.get("hora_ingreso"),
                    "hora_salida": row_data.get("hora_salida"),
                    "observaciones": row_data.get("observaciones") # Add observaciones
                }
                data["Compra"].append(compra_entry)
                logger.debug(f"Added Compra entry: {compra_entry}")
            elif entry_type == "Venta":
                 # Map Excel data to Venta model keys
                venta_entry = {
                    "id": row_data.get("registro_id"),
                    "cliente": row_data.get("contraparte"),
                    "mercaderia": row_data.get("producto"),
                    "bruto": row_data.get("peso_bruto"),
                    "tara": row_data.get("peso_tara"),
                    "merma": row_data.get("merma"),
                    "neto": row_data.get("peso_neto"),
                    "precio_kg": row_data.get("precio_x_kg"),
                    "importe": row_data.get("importe"),
                    "transporte": row_data.get("chofer/transporte"), # Map Transporte from Chofer/Transporte
                    "patente": row_data.get("patente"),
                    "incoterm": row_data.get("incoterm"),
                    "fecha": row_data.get("fecha_operacion"),
                    "hora_ingreso": row_data.get("hora_ingreso"),
                    "hora_salida": row_data.get("hora_salida"),
                    "remito": row_data.get("remito"),
                    "observaciones": row_data.get("observaciones") # Add observaciones
                }
                data["Venta"].append(venta_entry)
                logger.debug(f"Added Venta entry: {venta_entry}")

        logger.info(f"Successfully loaded {len(data['Compra'])} Compra entries and {len(data['Venta'])} Venta entries.")

    except Exception as e:
        logger.error(f"Error loading daily data from Excel: {e}", exc_info=True)

    return data

# --- No main execution block ---
